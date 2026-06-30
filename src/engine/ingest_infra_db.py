"""
src/engine/ingest_infra_db.py — INFRA_SNCF Excel Database Ingester (Layer 2)
=============================================================================
Reads the SNCF infrastructure database Excel file (LPK_752000_534000_Db_V2)
and extracts structured data for the digital twin.

Architecture Position (Layer 2 — Data Ingestion):
    - READS:   data/New_data/INFRA_SNCF/LPK_752000_534000_Db_V2_*.xlsx
    - WRITES:  data/processed/infra_sncf_metadata.json  (site metadata)
               data/processed/infra_sncf_rainfall.json  (hourly rainfall history)
               data/processed/infra_sncf_thresholds.json (alert & vigilance thresholds)
    - FEEDS:   src/engine/swi_calculator.py  (rainfall return periods)
               src/dashboard/app_main.py     (threshold display)

Excel Sheets (8 total):
    Fiche:           Site identification (Ligne, PK, PRI, Classement)
    Data:            28,028 rows of hourly rainfall intensity history (1997-present)
    Seuils:          Alert & vigilance rainfall thresholds per accumulation window
    Shyreg:          Return period statistics (10yr, 100yr) per accumulation window
    Incidents:       Historical incident records (failures, damage)
    EvolutionSeuil:  Threshold change log
    Graphs, RefGraphs: Chart configuration (not data)

Authors: TRAN Trong-Tin (Antigravity-generated)
Project: SNCF Railway Flood-Risk Digital Twin (Master Capstone)
"""

import json
import logging
from pathlib import Path
from typing import Dict, Optional, Union

import openpyxl

logger = logging.getLogger(__name__)


class INFRASNCFIngester:
    """Reads and structures the SNCF infrastructure database Excel file."""

    def __init__(self, xlsx_path: Union[str, Path]):
        self._path = Path(xlsx_path)
        self._wb = None

    def open(self):
        """Open the Excel workbook (read-only mode)."""
        if not self._path.exists():
            raise FileNotFoundError(f"Excel file not found: {self._path}")
        self._wb = openpyxl.load_workbook(str(self._path), read_only=True, data_only=True)
        logger.info(f"Opened INFRA_SNCF database: {self._path.name}")
        logger.info(f"Sheets: {self._wb.sheetnames}")

    def close(self):
        if self._wb:
            self._wb.close()
            self._wb = None

    def __enter__(self):
        self.open()
        return self

    def __exit__(self, *args):
        self.close()

    # ------------------------------------------------------------------
    # Sheet: Fiche (Site Identification)
    # ------------------------------------------------------------------
    def extract_site_metadata(self) -> Dict:
        """Extract site identification from the 'Fiche' sheet."""
        ws = self._wb['Fiche']

        # Headers in row 1, data in row 2
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value) if cell.value else "")

        data = {}
        for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
            for i, val in enumerate(row):
                if i < len(headers) and headers[i]:
                    data[headers[i]] = str(val) if val is not None else ""

        result = {
            "ligne": data.get("Ligne", ""),
            "nom": data.get("Nom", ""),
            "pk_debut": data.get("PK_debut", ""),
            "pk_fin": data.get("PK_fin", ""),
            "pk_pluie": data.get("PK_pluie", ""),
            "pri": data.get("PRI", ""),
            "up": data.get("UP", ""),
            "classement_ot": data.get("Classement_OT", ""),
            "consigne_intemperie": data.get("Consigne_Intemperie", ""),
            "incident_redoute": data.get("Incident_redoute", ""),
            "priorite": data.get("Priorite_1_ou_2", ""),
            "raison_non_seuil": data.get("Raison_non_seuil", ""),
            "commentaire1": data.get("Commentaire1_Reunion", ""),
            "commentaire2": data.get("Commentaire2_Reunion_MessageAlertes", ""),
        }
        logger.info(f"Site: Ligne {result['ligne']}, {result['nom']}, "
                     f"PK {result['pk_debut']}-{result['pk_fin']}")
        return result

    # ------------------------------------------------------------------
    # Sheet: Data (Rainfall History)
    # ------------------------------------------------------------------
    def extract_rainfall_data(self) -> Dict:
        """Extract hourly rainfall intensity data from the 'Data' sheet.

        Returns:
            Dict with keys: 'headers', 'n_rows', 'accumulation_windows',
                            'data' (list of dicts per row)
        """
        ws = self._wb['Data']

        # Read headers
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value) if cell.value else "")
        logger.info(f"Data sheet headers: {headers}")

        # Accumulation windows: 30Min, 1H, 4H, 12H, 24H, 48H, 72H
        accum_windows = [h for h in headers if h in ('30Min', '1H', '4H', '12H', '24H', '48H', '72H')]

        # Read data rows
        rows = []
        site_info = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    key = headers[i]
                    if key and val is not None:
                        row_dict[key] = str(val)

            # Store site info from first row with data
            if row_idx == 0 and row_dict.get('Ligne'):
                site_info = {k: row_dict.get(k, '') for k in ['Ligne', 'Nom', 'PK_debut', 'PK_fin']}

            # Only store rows with Index (timestamp) data
            if 'Index' in row_dict and row_dict['Index']:
                entry = {
                    'timestamp': row_dict['Index'],
                }
                for w in accum_windows:
                    val_str = row_dict.get(w, '')
                    try:
                        entry[w] = float(val_str) if val_str else None
                    except ValueError:
                        entry[w] = None
                rows.append(entry)

        result = {
            'site': site_info,
            'accumulation_windows': accum_windows,
            'n_rows': len(rows),
            'data': rows,
        }
        logger.info(f"Rainfall data: {len(rows)} rows, windows: {accum_windows}")
        return result

    # ------------------------------------------------------------------
    # Sheet: Seuils (Alert Thresholds)
    # ------------------------------------------------------------------
    def extract_thresholds(self) -> Dict:
        """Extract rainfall alert and vigilance thresholds from 'Seuils' sheet."""
        ws = self._wb['Seuils']

        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value) if cell.value else "")

        thresholds = {}
        # Only 1 data row expected
        for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
            for i, val in enumerate(row):
                if i < len(headers) and headers[i]:
                    thresholds[headers[i]] = str(val) if val is not None else ""

        result = {
            'periode_pluie': thresholds.get('periode_pluie', ''),
            'seuil_alerte_mm': thresholds.get('Seuil_Alerte', ''),
            'seuil_vigilance_mm': thresholds.get('Seuil_Vigilance', ''),
            'valeur_test': thresholds.get('Valeur_test', ''),
            'valeur_test2': thresholds.get('Valeur_test2', ''),
        }
        logger.info(f"Thresholds: Alert={result['seuil_alerte_mm']}mm, "
                     f"Vigilance={result['seuil_vigilance_mm']}mm")
        return result

    # ------------------------------------------------------------------
    # Sheet: Shyreg (Return Period Statistics)
    # ------------------------------------------------------------------
    def extract_return_periods(self) -> Dict:
        """Extract Shyreg return period statistics."""
        ws = self._wb['Shyreg']

        headers_row1 = []
        for cell in ws[1]:
            headers_row1.append(str(cell.value) if cell.value else "")

        # Row 2 has the sub-headers (30MIN, 1H, 4H, etc.)
        sub_headers = []
        for cell in ws[2]:
            sub_headers.append(str(cell.value) if cell.value else "")

        result = {
            'headers': headers_row1,
            'sub_headers': sub_headers,
        }

        # The actual data structure varies; extract what we can
        logger.info(f"Shyreg headers: {headers_row1[:10]}")
        logger.info(f"Shyreg sub-headers: {sub_headers[:10]}")
        return result

    # ------------------------------------------------------------------
    # Sheet: Incidents (Historical Events)
    # ------------------------------------------------------------------
    def extract_incidents(self) -> list:
        """Extract historical incident records."""
        ws = self._wb['Incidents']

        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value) if cell.value else "")

        incidents = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            has_data = False
            for i, val in enumerate(row):
                if i < len(headers) and headers[i] and val is not None:
                    row_dict[headers[i]] = str(val)
                    has_data = True
            if has_data and any(v for v in row_dict.values()):
                incidents.append(row_dict)

        logger.info(f"Incidents: {len(incidents)} records")
        return incidents

    # ------------------------------------------------------------------
    # Full extraction
    # ------------------------------------------------------------------
    def extract_all(self, output_dir: Union[str, Path]) -> Dict:
        """Extract all sheets and save to JSON files."""
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        # Site metadata
        metadata = self.extract_site_metadata()
        with open(output_dir / "infra_sncf_metadata.json", 'w', encoding='utf-8') as f:
            json.dump(metadata, f, indent=2, ensure_ascii=False)

        # Rainfall data (large — save without indent for size)
        rainfall = self.extract_rainfall_data()
        with open(output_dir / "infra_sncf_rainfall.json", 'w', encoding='utf-8') as f:
            json.dump(rainfall, f, ensure_ascii=False)
        logger.info(f"Saved rainfall data: {rainfall['n_rows']} rows")

        # Thresholds
        thresholds = self.extract_thresholds()
        with open(output_dir / "infra_sncf_thresholds.json", 'w', encoding='utf-8') as f:
            json.dump(thresholds, f, indent=2, ensure_ascii=False)

        # Return periods
        return_periods = self.extract_return_periods()
        with open(output_dir / "infra_sncf_shyreg.json", 'w', encoding='utf-8') as f:
            json.dump(return_periods, f, indent=2, ensure_ascii=False)

        # Incidents
        incidents = self.extract_incidents()
        with open(output_dir / "infra_sncf_incidents.json", 'w', encoding='utf-8') as f:
            json.dump(incidents, f, indent=2, ensure_ascii=False)

        summary = {
            'site': metadata,
            'rainfall_rows': rainfall['n_rows'],
            'accumulation_windows': rainfall['accumulation_windows'],
            'thresholds': thresholds,
            'n_incidents': len(incidents),
        }
        logger.info(f"Full extraction complete. Files saved to {output_dir}")
        return summary


# ======================================================================
# Standalone execution
# ======================================================================
if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s [%(name)s] %(message)s")

    from pathlib import Path
    ROOT = Path(r"c:\Users\ktstr\Documents\railway-flood-twin")
    xlsx = ROOT / "data" / "New_data" / "INFRA_SNCF" / \
           "LPK_752000_534000_Db_V2_South_Head_Tartaiguille_H_JN_LGV.xlsx"
    output_dir = ROOT / "data" / "processed"

    with INFRASNCFIngester(xlsx) as ingester:
        summary = ingester.extract_all(output_dir)

    print("\n=== Extraction Summary ===")
    print(f"  Site: Ligne {summary['site']['ligne']}, {summary['site']['nom']}")
    print(f"  PK: {summary['site']['pk_debut']} to {summary['site']['pk_fin']}")
    print(f"  Rainfall rows: {summary['rainfall_rows']}")
    print(f"  Accumulation windows: {summary['accumulation_windows']}")
    print(f"  Thresholds: Alert={summary['thresholds']['seuil_alerte_mm']}mm, "
          f"Vigilance={summary['thresholds']['seuil_vigilance_mm']}mm")
    print(f"  Incidents: {summary['n_incidents']} records")
    print("\nDone!")
